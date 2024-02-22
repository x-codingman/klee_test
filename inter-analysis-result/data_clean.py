import re
import openpyxl
from openpyxl.utils import get_column_letter

# Assuming the format of each line in the text file is:
# "KLEE: DETECT KERNEL MEMORY TAMPERING: x with y"
pattern = r"KLEE:  DETECT KERNEL MEMORY TAMPERING: (\w+) with (\w+)"
# Read the file and extract pairs
pairs = []
with open('/home/klee/klee_test/inter-analysis-result/test-info-inter-analysis/test_info.txt', 'r') as file:
    for line in file:
        match = re.search(pattern, line)
        if match:
            pairs.append(match.groups())

# Deduplicate and sort the function names
functions = sorted(set([func for pair in pairs for func in pair]))
print(functions)
# Create a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Fill the first row and column with function names
for index, func in enumerate(functions, start=1):
    ws.cell(row=index+1, column=1, value=func)
    ws.cell(row=1, column=index+1, value=func)

# Mark the pairs with a check mark
for x, y in pairs:
    x_index = functions.index(x) + 2
    y_index = functions.index(y) + 2
    ws.cell(row=x_index, column=y_index, value='✓')

# Set column width for better readability
for col in ws.columns:
    max_length = 0
    column = col[0].column # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

# Save the workbook
wb.save('output.xlsx')
