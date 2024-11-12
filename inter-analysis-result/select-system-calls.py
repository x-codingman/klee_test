import pandas as pd
from openpyxl import load_workbook
import openpyxl

class SystemCall:
    def __init__(self, name, call_type, accessible_areas, condition_num=0, condition_areas=None, accessible_num=0):
        self.name = name
        self.call_type = call_type
        self.condition_num = condition_num
        self.condition_num_max = condition_num
        self.condition_num_min = condition_num
        self.accessible_num = accessible_num
        self.accessible_areas = accessible_areas
        self.condition_areas = condition_areas if condition_areas else []

    def merge_accessible_areas(self, new_areas):
        self.accessible_areas = list(set(self.accessible_areas + new_areas))

    def __str__(self):
        return (f"System Call: {self.name}, Type: {self.call_type}, "
                f"Accessible Areas: {sorted(self.accessible_areas)}, "
                f"Condition Areas: {sorted(self.condition_areas)}")

class SystemCallManager:
    def __init__(self):
        self.calls = []

    def add_system_call(self, system_call):
        existing_call = self.find_system_call(system_call.name)
        if existing_call:
            existing_call.merge_accessible_areas(system_call.accessible_areas)
            print(existing_call.condition_num)
            # pick the maxnium 
            if existing_call.condition_num_min > system_call.condition_num:
               existing_call.condition_num_min = system_call.condition_num
            if existing_call.condition_num_max < system_call.condition_num:
               existing_call.condition_num_max = system_call.condition_num
            if(len(existing_call.accessible_areas)!=0):
                existing_call.accessible_areas.sort()
                existing_call.accessible_num = len(existing_call.accessible_areas)
        else:
            self.calls.append(system_call)

    def find_system_call(self, name):
        for call in self.calls:
            if call.name == name:
                return call
        return None
    
    def sort_calls(self):
        # 对calls进行排序，名字中包含"create"的排在后面
        self.calls.sort(key=lambda call: 'create' in call.name.lower())

    def export_to_xlsx(self, filename):
        
        data = [
            {
                'Name': call.name,
                'Type': call.call_type,
                'Accessible Areas': ','.join(map(str, call.accessible_areas)),
                'Accessible Num': int(call.accessible_num),
                'Condition Num': str(int(call.condition_num_max))+'/'+str(int(call.condition_num_min)),
                'Condition Areas': ','.join(map(str, call.condition_areas))
            }
            for call in self.calls
        ]
        df = pd.DataFrame(data)
        df.to_excel(filename, index=False, engine='openpyxl')

    def import_from_xlsx(self, filename):
        df = pd.read_excel(filename, engine='openpyxl')
        for _, row in df.iterrows():
            # Accessible Areas
            areas = list(map(int, row['Accessible Areas'].split(','))) if pd.notna(row['Accessible Areas']) else []

            # Condition Areas
            if pd.notna(row['Condition Areas']):
                if isinstance(row['Condition Areas'], str):
                    condition_areas = list(map(int, row['Condition Areas'].split(',')))
                elif isinstance(row['Condition Areas'], int):
                    condition_areas = [row['Condition Areas']]
                else:
                    condition_areas = []
            else:
                condition_areas = []

            # Create a new SystemCall object
            system_call = SystemCall(row['Name'], row['Type'], areas, condition_areas)
            self.add_system_call(system_call)

    def display_all_calls(self):
        for call in self.calls:
            if 'create' in call.name.lower():
                call.condition_num += 1
            print(call)



def find_path(manager, target_distance):
    solutions = []
    start_calls = [call for call in manager.calls if 'create' in call.name.lower()]
    manager.sort_calls()
    def backtrack(current_distance, path):
        if current_distance == target_distance:
            solutions.append(path.copy())
            return True
        if current_distance > target_distance:
            return False

        last_call, last_step = path[-1] if path else (None, None)
        
        for call in manager.calls:
            if last_call:
                all_conditions_matched = all((condition + last_step) in last_call.accessible_areas for condition in call.condition_areas)
                if not all_conditions_matched:
                    continue

            for step in call.accessible_areas:
                new_distance = current_distance + step
                if new_distance <= target_distance:
                    path.append((call, step))
                    if backtrack(new_distance, path):
                        return True
                    path.pop()
        return False

    for start_call in start_calls:
        for start_step in start_call.accessible_areas:
            if start_step <= target_distance:
                if backtrack(start_step, [(start_call, start_step)]):
                    break

    return solutions


def find_path_all(manager, target_distance):
    solutions = []
    start_calls = [call for call in manager.calls if 'create' in call.name.lower()]

    def backtrack(current_distance, path):
        if current_distance == target_distance:
            solutions.append(path.copy())
            return
        if current_distance > target_distance:
            return

        last_call, last_step = path[-1] if path else (None, None)
        
        for call in manager.calls:
            if last_call:  # 确保这不是第一次迭代，因为第一次迭代没有last_call
                all_conditions_matched = all((condition + last_step) in last_call.accessible_areas for condition in call.condition_areas)
                if not all_conditions_matched:
                    continue

            for step in call.accessible_areas:
                new_distance = current_distance + step
                if new_distance <= target_distance:
                    path.append((call, step))
                    backtrack(new_distance, path)
                    path.pop()

    for start_call in start_calls:
        for start_step in start_call.accessible_areas:
            if start_step <= target_distance:
                backtrack(start_step, [(start_call, start_step)])

    return solutions





manager = SystemCallManager()
manager_can_be_forged = SystemCallManager()
manager_all = SystemCallManager()

workbook = load_workbook(filename="writable_location_each_syscall_analysis.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题
    syscall_name, object_type, mo_name, offset, value_controllable, can_be_forged_id, condition_num = row[1], row[2], row[3], row[6], row[7], row[8],row[9]
    
    # if mo_name == "lazy_alloc1":
    #     if condition_num == False:
    #         syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
    #     else:
    #         syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))

    ## 找到value controllable的
    if mo_name == "lazy_alloc1" and value_controllable == True:
        if condition_num == False:
            syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
        else:
            syscall = manager.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))
    if mo_name == "lazy_alloc1" and value_controllable == True and can_be_forged_id:
        if condition_num == False:
            syscall = manager_can_be_forged.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
        else:
            syscall = manager_can_be_forged.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))
    if mo_name == "lazy_alloc1":
        if condition_num == False:
            syscall = manager_all.add_system_call(SystemCall(syscall_name, object_type, [offset], 0))
        else:
            syscall = manager_all.add_system_call(SystemCall(syscall_name, object_type, [offset], condition_num))
manager.display_all_calls()
manager_all.export_to_xlsx('vulnerable_system_calls_all_v2.xlsx')
manager.export_to_xlsx('vulnerable_system_calls_value_controllable_v2.xlsx')
manager_can_be_forged.export_to_xlsx('vulnerable_system_calls_value_controllable_and_can_be_forged_id_v2.xlsx')
# new_manager = SystemCallManager()
# new_manager.import_from_xlsx('vulnerable_system_calls.xlsx')    
# new_manager.display_all_calls()



# target_distance = 256  # 目标距离
# paths = find_path_all(new_manager, target_distance)
# min_length = 100
# paths_min = []
# if(len(paths) == 0):
#         print("No result for distance of" + str(target_distance))
# for path in paths:
#     length = len(path)
#     if length < min_length:
#         paths_min = []
#         paths_min.append(path)
#         print(" -> ".join(f"{call.name} (step {step})" for call, step in path))
#         min_length = length
#     elif length == min_length:
#         paths_min.append(path)
# print(len(paths_min))
# for path in paths_min:
#     print(" -> ".join(f"{call.name} (step {step})" for call, step in path))
# for i in range(100):
#     distance = i*8
#     paths = find_path(new_manager, distance)
#     if(len(paths) == 0):
#         print("No result for distance of" + str(distance))
    # for path in paths:
    #     print(" -> ".join(f"{call.name} (step {step})" for call, step in path))


# Print the information of each syscall
#print_syscalls(syscalls)
